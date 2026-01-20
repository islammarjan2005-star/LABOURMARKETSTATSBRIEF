#!/usr/bin/env python3
"""
Creates an Excel template with placeholders from the source workbook.
Replaces calculated values with {{PLACEHOLDER_NAME}} markers.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy
import shutil

# Source and output paths
SOURCE_FILE = "December 25 LM Statss.xlsx"
TEMPLATE_FILE = "LM_Stats_Template.xlsx"

# Placeholder styling - yellow highlight to make them visible
PLACEHOLDER_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PLACEHOLDER_FONT = Font(bold=True, color="0000FF")

# ==============================================================================
# DASHBOARD PLACEHOLDERS (Sheet: Dashboard)
# ==============================================================================
# Format: (row, col, placeholder_name)
# Based on the structure: row 6-17, cols D-G (4-7) for values

DASHBOARD_PLACEHOLDERS = [
    # Row 6: Employment (000s) 16+
    (6, 4, "EMP16_CUR"),
    (6, 5, "EMP16_DQ"),
    (6, 6, "EMP16_DY"),
    (6, 7, "EMP16_DC"),

    # Row 7: Employment rate (16-64)
    (7, 4, "EMP_RT_CUR"),
    (7, 5, "EMP_RT_DQ"),
    (7, 6, "EMP_RT_DY"),
    (7, 7, "EMP_RT_DC"),

    # Row 8: Unemployment, 16+ (000s)
    (8, 4, "UNEMP16_CUR"),
    (8, 5, "UNEMP16_DQ"),
    (8, 6, "UNEMP16_DY"),
    (8, 7, "UNEMP16_DC"),

    # Row 9: Unemployment rate, 16+
    (9, 4, "UNEMP_RT_CUR"),
    (9, 5, "UNEMP_RT_DQ"),
    (9, 6, "UNEMP_RT_DY"),
    (9, 7, "UNEMP_RT_DC"),

    # Row 10: Economic inactivity (000s) (16-64)
    (10, 4, "INACT_CUR"),
    (10, 5, "INACT_DQ"),
    (10, 6, "INACT_DY"),
    (10, 7, "INACT_DC"),

    # Row 11: Economic inactivity rate (16-64)
    (11, 4, "INACT_RT_CUR"),
    (11, 5, "INACT_RT_DQ"),
    (11, 6, "INACT_RT_DY"),
    (11, 7, "INACT_RT_DC"),

    # Row 12: 50-64s inactivity (000s)
    (12, 4, "INACT5064_CUR"),
    (12, 5, "INACT5064_DQ"),
    (12, 6, "INACT5064_DY"),
    (12, 7, "INACT5064_DC"),

    # Row 13: 50-64s inactivity rate
    (13, 4, "INACT5064_RT_CUR"),
    (13, 5, "INACT5064_RT_DQ"),
    (13, 6, "INACT5064_RT_DY"),
    (13, 7, "INACT5064_RT_DC"),

    # Row 14: Payroll employees (000s)
    (14, 4, "PAYROLL_CUR"),
    (14, 5, "PAYROLL_DQ"),
    (14, 6, "PAYROLL_DY"),
    (14, 7, "PAYROLL_DC"),

    # Row 15: Vacancies (000s)
    (15, 4, "VAC_CUR"),
    (15, 5, "VAC_DQ"),
    (15, 6, "VAC_DY"),
    (15, 7, "VAC_DC"),

    # Row 16: Wage growth cash terms
    (16, 4, "WAGES_CUR"),
    (16, 5, "WAGES_DQ"),
    (16, 6, "WAGES_DY"),
    (16, 7, "WAGES_DC"),

    # Row 17: Wage growth adjusted for inflation
    (17, 4, "WAGES_CPI_CUR"),
    (17, 5, "WAGES_CPI_DQ"),
    (17, 6, "WAGES_CPI_DY"),
    (17, 7, "WAGES_CPI_DC"),
]

# ==============================================================================
# SHEET 2 PLACEHOLDERS (LFS Summary)
# ==============================================================================
# Structure: Rows 5-10 contain summary data
# Columns: B=Emp level, C=Emp rate, D=Unemp level, E=Unemp rate,
#          F=Activity level, G=Activity rate, H=Inact level, I=Inact rate

SHEET2_PLACEHOLDERS = [
    # Row 5: Current
    (5, 2, "EMP16_CUR"),
    (5, 3, "EMP_RT_CUR"),
    (5, 4, "UNEMP16_CUR"),
    (5, 5, "UNEMP_RT_CUR"),
    # Activity = Emp + Unemp (formula), skip
    # (5, 6, "ACTIVITY_CUR"),
    # (5, 7, "ACTIVITY_RT_CUR"),
    (5, 8, "INACT_CUR"),
    (5, 9, "INACT_RT_CUR"),

    # Row 6: Quarterly change
    (6, 2, "EMP16_DQ"),
    (6, 3, "EMP_RT_DQ"),
    (6, 4, "UNEMP16_DQ"),
    (6, 5, "UNEMP_RT_DQ"),
    (6, 8, "INACT_DQ"),
    (6, 9, "INACT_RT_DQ"),

    # Row 7: Change year on year
    (7, 2, "EMP16_DY"),
    (7, 3, "EMP_RT_DY"),
    (7, 4, "UNEMP16_DY"),
    (7, 5, "UNEMP_RT_DY"),
    (7, 8, "INACT_DY"),
    (7, 9, "INACT_RT_DY"),

    # Row 8: Change since Covid
    (8, 2, "EMP16_DC"),
    (8, 3, "EMP_RT_DC"),
    (8, 4, "UNEMP16_DC"),
    (8, 5, "UNEMP_RT_DC"),
    (8, 8, "INACT_DC"),
    (8, 9, "INACT_RT_DC"),

    # Row 10: Change since 2024 election
    (10, 2, "EMP16_DE"),
    (10, 3, "EMP_RT_DE"),
    (10, 4, "UNEMP16_DE"),
    (10, 5, "UNEMP_RT_DE"),
    (10, 8, "INACT_DE"),
    (10, 9, "INACT_RT_DE"),
]

# ==============================================================================
# PAYROLL SHEET PLACEHOLDERS (Sheet: 1. Payrolled employees (UK))
# ==============================================================================

PAYROLL_PLACEHOLDERS = [
    # Row 2: Summary row
    (2, 2, "PAYROLL_CUR"),      # Current
    (2, 4, "PAYROLL_DQ"),       # Change on quarter
    (2, 5, "PAYROLL_DY"),       # Change year on year
    (2, 6, "PAYROLL_DC"),       # Change since Covid
    (2, 7, "PAYROLL_DE"),       # Change since election
]

# ==============================================================================
# SHEET CONFIGURATION
# ==============================================================================

SHEETS_CONFIG = {
    "Dashboard": {
        "placeholders": DASHBOARD_PLACEHOLDERS,
        "title_cell": (5, 2),  # Where period label goes
        "title_placeholder": "LFS_PERIOD"
    },
    "2": {
        "placeholders": SHEET2_PLACEHOLDERS,
        "source_data_start": 16,  # Row where source data begins
    },
    "1. Payrolled employees (UK)": {
        "placeholders": PAYROLL_PLACEHOLDERS,
        "source_data_start": 11,
    },
}


def insert_placeholder(ws, row, col, placeholder_name, style=True):
    """Insert a placeholder into a cell."""
    cell = ws.cell(row=row, column=col)
    cell.value = f"{{{{{placeholder_name}}}}}"
    if style:
        cell.fill = PLACEHOLDER_FILL
        cell.font = PLACEHOLDER_FONT


def clear_source_data(ws, start_row, max_col=None):
    """Clear all data from start_row to the end of the sheet."""
    if max_col is None:
        max_col = ws.max_column

    cleared = 0
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None
                cleared += 1
    return cleared


# Define where source data starts for each sheet (to be cleared)
SOURCE_DATA_CONFIG = {
    "Dashboard": None,  # No source data to clear
    "2": {"start_row": 18, "max_col": 66},  # LFS source data
    "1. Payrolled employees (UK)": {"start_row": 11, "max_col": 19},
    "23. Employees Industry": {"start_row": 14, "max_col": 44},
    "5": {"start_row": 16, "max_col": 18},
    "10": {"start_row": 10, "max_col": 18},
    "11": {"start_row": 10, "max_col": 50},
    "13": {"start_row": 18, "max_col": 29},
    "15": {"start_row": 18, "max_col": 28},
    "18": {"start_row": 10, "max_col": 14},
    "20": {"start_row": 10, "max_col": 17},
    "21": {"start_row": 10, "max_col": 28},
    "AWE Real_CPI": {"start_row": 17, "max_col": 16},
    "1a": {"start_row": 10, "max_col": 18},
    "1b": {"start_row": 10, "max_col": 13},
}


def create_template():
    """Create the template file with placeholders and cleared source data."""
    print(f"Loading source file: {SOURCE_FILE}")
    wb = openpyxl.load_workbook(SOURCE_FILE)

    # First, insert placeholders in configured sheets
    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: Sheet '{sheet_name}' not found, skipping")
            continue

        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]

        # Insert placeholders
        placeholders = config.get("placeholders", [])
        for row, col, ph_name in placeholders:
            insert_placeholder(ws, row, col, ph_name)
            print(f"  Inserted {{{{{ph_name}}}}} at R{row}C{col}")

        # Insert title placeholder if specified
        if "title_cell" in config and "title_placeholder" in config:
            tr, tc = config["title_cell"]
            insert_placeholder(ws, tr, tc, config["title_placeholder"])
            print(f"  Inserted title placeholder at R{tr}C{tc}")

    # Now clear source data from all sheets
    print("\n=== Clearing source data ===")
    for sheet_name, clear_config in SOURCE_DATA_CONFIG.items():
        if clear_config is None:
            continue
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        start_row = clear_config["start_row"]
        max_col = clear_config.get("max_col", ws.max_column)

        cleared = clear_source_data(ws, start_row, max_col)
        print(f"  {sheet_name}: Cleared {cleared} cells from row {start_row}")

    # Save the template
    print(f"\nSaving template to: {TEMPLATE_FILE}")
    wb.save(TEMPLATE_FILE)
    print("Done!")

    return TEMPLATE_FILE


if __name__ == "__main__":
    create_template()
